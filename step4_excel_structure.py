from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_structure():
    """Create basic Excel structure with summary and navigation tabs."""
    
    print("Creating Excel structure...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Tab 1: Summary Statistics
    print("Creating summary tab...")
    ws_summary = wb.create_sheet("Summary_Statistics")
    ws_summary['A1'] = "OVERLAP ANALYSIS SUMMARY"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    # Add summary data
    for r in dataframe_to_rows(summary_stats, index=False, header=True):
        ws_summary.append(r)
    
    # Tab 2: Navigation Index
    print("Creating navigation tab...")
    ws_nav = wb.create_sheet("Navigation_Index")
    ws_nav['A1'] = "OVERLAPPING PROVIDER ANALYSIS INDEX"
    ws_nav['A1'].font = Font(size=16, bold=True)
    
    # Add navigation headers
    headers = ['Case #', 'Provider Name', 'Label A', 'Label B', 'Overlap Score', 'Analysis Link']
    for col, header in enumerate(headers, 1):
        ws_nav.cell(row=3, column=col, value=header).font = Font(bold=True)
    
    # Add navigation data with hyperlinks
    for idx, row in navigation_df.iterrows():
        ws_nav[f'A{idx+4}'] = row['Case_Number']
        ws_nav[f'B{idx+4}'] = row['Provider_Name']
        ws_nav[f'C{idx+4}'] = row['Label_A']
        ws_nav[f'D{idx+4}'] = row['Label_B']
        ws_nav[f'E{idx+4}'] = row['Overlap_Score']
        ws_nav[f'F{idx+4}'] = f"→ {row['Tab_Name']}"
        ws_nav[f'F{idx+4}'].hyperlink = f"#{row['Tab_Name']}!A1"
        ws_nav[f'F{idx+4}'].font = Font(color='0000FF', underline='single')
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_nav]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    print("✅ Basic Excel structure created")
    return wb

# Execute Step 4
wb = create_excel_structure()

# Save basic structure for recovery
recovery_filename = f"RECOVERY_{EXPORT_FILENAME}"
wb.save(recovery_filename)
print(f"✅ Step 4 Complete: Basic structure saved to {recovery_filename}")