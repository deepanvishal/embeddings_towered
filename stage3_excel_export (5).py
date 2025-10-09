import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import datetime
import gc
import os
from sklearn.decomposition import PCA

# Configuration - Easy to change for testing vs full run
EXPORT_LIMIT = 5  # Change to None for all providers, or any number for testing
EXPORT_FILENAME = f"Overlapping_Provider_Analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

print(f"Excel Export Configuration:")
print(f"Export Limit: {EXPORT_LIMIT if EXPORT_LIMIT else 'All providers'}")
print(f"Output File: {EXPORT_FILENAME}")

# PERFORMANCE OPTIMIZATION: Pre-compute expensive operations
print("Pre-computing data for performance...")

# Pre-compute all provider analysis data once
def precompute_all_analysis_data():
    """Pre-compute all analysis data to speed up Excel generation."""
    
    print("Pre-computing provider analysis data...")
    export_cases = analysis_df.nlargest(EXPORT_LIMIT if EXPORT_LIMIT else len(analysis_df), 'overlap_score')
    
    precomputed_data = {}
    
    for idx, (_, row) in enumerate(export_cases.iterrows(), 1):
        print(f"Pre-computing case {idx}/{len(export_cases)}: {row['outlier_name']}")
        
        pin_outlier = row['outlier_pin']
        pin_same = row['typical_same_pin']
        pin_other = row['typical_other_pin']
        
        # Pre-compute all data for this case
        case_data = {
            'procedures': {
                'outlier': get_provider_codes(pin_outlier, 'Procedure', 10),
                'typical_same': get_provider_codes(pin_same, 'Procedure', 10),
                'typical_other': get_provider_codes(pin_other, 'Procedure', 10),
                'common_same': find_common_codes(pin_outlier, pin_same, 'Procedure', 10),
                'common_other': find_common_codes(pin_outlier, pin_other, 'Procedure', 10)
            },
            'diagnoses': {
                'outlier': get_provider_codes(pin_outlier, 'Diagnosis', 10),
                'typical_same': get_provider_codes(pin_same, 'Diagnosis', 10),
                'typical_other': get_provider_codes(pin_other, 'Diagnosis', 10),
                'common_same': find_common_codes(pin_outlier, pin_same, 'Diagnosis', 10),
                'common_other': find_common_codes(pin_outlier, pin_other, 'Diagnosis', 10)
            },
            'demographics': {
                'outlier': get_provider_demographics(pin_outlier),
                'typical_same': get_provider_demographics(pin_same),
                'typical_other': get_provider_demographics(pin_other)
            },
            'costs': {
                'outlier': get_provider_costs(pin_outlier),
                'typical_same': get_provider_costs(pin_same),
                'typical_other': get_provider_costs(pin_other)
            }
        }
        
        # Pre-compute comparison tables
        case_data['demographics']['comparison_same'] = create_comparison_table(
            case_data['demographics']['outlier'], case_data['demographics']['typical_same'],
            row['outlier_name'], row['typical_same_name'], 'demo'
        )
        case_data['demographics']['comparison_other'] = create_comparison_table(
            case_data['demographics']['outlier'], case_data['demographics']['typical_other'],
            row['outlier_name'], row['typical_other_name'], 'demo'
        )
        case_data['costs']['comparison_same'] = create_comparison_table(
            case_data['costs']['outlier'], case_data['costs']['typical_same'],
            row['outlier_name'], row['typical_same_name'], 'cost'
        )
        case_data['costs']['comparison_other'] = create_comparison_table(
            case_data['costs']['outlier'], case_data['costs']['typical_other'],
            row['outlier_name'], row['typical_other_name'], 'cost'
        )
        
        precomputed_data[pin_outlier] = case_data
    
    print("Pre-computation complete!")
    return export_cases, precomputed_data

def clean_tab_name(name):
    """Clean provider name for Excel tab naming."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?', '<', '>', '|']
    clean_name = str(name)
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    return clean_name[:25]  # Limit to 25 characters for safety

def prepare_export_data(export_cases):
    """Prepare navigation data."""
    
    summary_stats = overlap_summary.copy()
    
    navigation_index = []
    for idx, (_, row) in enumerate(export_cases.iterrows(), 1):
        clean_name = clean_tab_name(row['outlier_name'])
        navigation_index.append({
            'Case_Number': idx,
            'Provider_Name': row['outlier_name'],
            'Label_A': row['outlier_label'].title(),
            'Label_B': row['overlap_label'].title(), 
            'Overlap_Score': round(row['overlap_score'], 3),
            'Tab_Name': f"Case_{idx}_{clean_name}"
        })
    
    navigation_df = pd.DataFrame(navigation_index)
    return summary_stats, navigation_df

def create_embedding_comparison_chart(outlier_pin, typical_same_pin, typical_other_pin, outlier_name, typical_same_name, typical_other_name):
    """Create embedding comparison chart."""
    
    outlier_idx = df[df['PIN'] == outlier_pin].index[0]
    typical_same_idx = df[df['PIN'] == typical_same_pin].index[0] 
    typical_other_idx = df[df['PIN'] == typical_other_pin].index[0]
    
    outlier_emb = embeddings[outlier_idx - df.index[0]]
    typical_same_emb = embeddings[typical_same_idx - df.index[0]]
    typical_other_emb = embeddings[typical_other_idx - df.index[0]]
    
    plt.figure(figsize=(16, 8))
    
    x_positions = range(64)
    plt.plot(x_positions, outlier_emb, marker='o', linewidth=2, markersize=4, 
             label=f'{outlier_name} (Overlapping Provider)', color='red', alpha=0.8)
    plt.plot(x_positions, typical_same_emb, marker='s', linewidth=2, markersize=4,
             label=f'{typical_same_name} (Same Group)', color='blue', alpha=0.8)
    plt.plot(x_positions, typical_other_emb, marker='^', linewidth=2, markersize=4,
             label=f'{typical_other_name} (Other Group)', color='green', alpha=0.8)
    
    plt.xlabel('Embedding Dimension', fontsize=12)
    plt.ylabel('Embedding Value', fontsize=12)
    plt.title('Provider Embedding Comparison: 64-Dimensional Profile', fontsize=14, fontweight='bold')
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    temp_filename = f"temp_embedding_{outlier_pin}.png"
    plt.savefig(temp_filename, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    
    return temp_filename

def create_pca_chart(outlier_pin, typical_same_pin, typical_other_pin, outlier_label, overlap_label, outlier_name, typical_same_name, typical_other_name):
    """Create PCA visualization chart."""
    
    same_mask = df['label'] == outlier_label
    other_mask = df['label'] == overlap_label
    
    group_embeddings = np.vstack([
        embeddings[same_mask.values],
        embeddings[other_mask.values]
    ])
    
    pca = PCA(n_components=2, random_state=42)
    pca_coords = pca.fit_transform(group_embeddings)
    
    same_count = same_mask.sum()
    pca_same = pca_coords[:same_count]
    pca_other = pca_coords[same_count:]
    
    same_centroid_emb = embeddings[same_mask.values].mean(axis=0)
    other_centroid_emb = embeddings[other_mask.values].mean(axis=0)
    
    centroids_pca = pca.transform([same_centroid_emb, other_centroid_emb])
    same_centroid_pca = centroids_pca[0]
    other_centroid_pca = centroids_pca[1]
    
    plt.figure(figsize=(14, 8))
    
    plt.scatter(pca_same[:, 0], pca_same[:, 1], alpha=0.6, label=outlier_label.title(), s=30, color='lightblue')
    plt.scatter(pca_other[:, 0], pca_other[:, 1], alpha=0.6, label=overlap_label.title(), s=30, color='lightcoral')
    
    plt.scatter(same_centroid_pca[0], same_centroid_pca[1], marker='X', s=400, 
                color='blue', edgecolors='black', linewidth=2, label=f'{outlier_label.title()} Centroid')
    plt.scatter(other_centroid_pca[0], other_centroid_pca[1], marker='X', s=400, 
                color='red', edgecolors='black', linewidth=2, label=f'{overlap_label.title()} Centroid')
    
    key_providers = [
        (outlier_pin, '*', 'darkred', f'{outlier_name}', 300),
        (typical_same_pin, 'o', 'darkblue', f'{outlier_label.title()} - {typical_same_name}', 200),
        (typical_other_pin, 's', 'darkgreen', f'{overlap_label.title()} - {typical_other_name}', 200)
    ]
    
    for pin, marker, color, label_text, size in key_providers:
        provider_idx = df[df['PIN'] == pin].index
        if len(provider_idx) > 0:
            provider_emb = embeddings[provider_idx[0] - df.index[0]]
            provider_pca = pca.transform([provider_emb])[0]
            
            plt.scatter(provider_pca[0], provider_pca[1], marker=marker, s=size, 
                       color=color, edgecolors='black', linewidth=1, label=label_text)
            
            if pin == outlier_pin:
                plt.plot([provider_pca[0], same_centroid_pca[0]], 
                        [provider_pca[1], same_centroid_pca[1]], 
                        'b--', alpha=0.7, linewidth=2, label='Distance to Own Centroid')
                plt.plot([provider_pca[0], other_centroid_pca[0]], 
                        [provider_pca[1], other_centroid_pca[1]], 
                        'r--', alpha=0.7, linewidth=2, label='Distance to Other Centroid')
    
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.title(f'PCA Visualization: {outlier_label.title()} vs {overlap_label.title()}\n(Preserves True Centroid Distances)', fontsize=14, fontweight='bold')
    plt.xlabel(f'PC1 ({pca.explained_variance_ratio_[0]:.1%} variance)', fontsize=12)
    plt.ylabel(f'PC2 ({pca.explained_variance_ratio_[1]:.1%} variance)', fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    temp_filename = f"temp_pca_{outlier_pin}.png"
    plt.savefig(temp_filename, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    
    del group_embeddings, pca_coords, pca_same, pca_other
    gc.collect()
    
    return temp_filename

def add_back_link(ws, tab_name):
    """Add back to index link."""
    ws['A1'] = '← BACK TO INDEX'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, italic=True, color='0000FF', underline='single')
    ws['A1'].hyperlink = "#Navigation_Index!A1"

def style_worksheet(ws, title):
    """Apply professional styling."""
    title_font = Font(name='Arial', size=16, bold=True, color='1F497D')
    ws['A1'].font = title_font
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def write_analysis_to_worksheet(ws, selected_row, analysis_data, tab_name):
    """Write complete analysis with all sections and charts."""
    
    add_back_link(ws, tab_name)
    current_row = 3
    
    # Header information
    ws[f'A{current_row}'] = f"OVERLAPPING PROVIDER ANALYSIS: {selected_row['outlier_name']}"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    current_row += 2
    
    ws[f'A{current_row}'] = f"Selected Provider for Analysis: {selected_row['outlier_name']} ({selected_row['outlier_label'].title()})"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider embeddings indicate the provider is close to {selected_row['overlap_label'].title()}."
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['overlap_label'].title()} is {selected_row['typical_other_name']}"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['outlier_label'].title()} is {selected_row['typical_same_name']}"
    current_row += 2
    
    # Distance Metrics
    ws[f'A{current_row}'] = "Distance Metrics:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to own centroid ({selected_row['outlier_label'].title()}): {selected_row['own_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to overlapping centroid ({selected_row['overlap_label'].title()}): {selected_row['other_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Overlap coefficient: {selected_row['overlap_score']:.4f}"
    current_row += 3
    
    # WITHIN-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"WITHIN-GROUP ANALYSIS: {selected_row['outlier_label'].title()} Providers"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Procedures - Outlier
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No procedure data available"
        current_row += 1
    current_row += 1
    
    # Procedures - Typical Same
    ws[f'A{current_row}'] = f"{selected_row['typical_same_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['typical_same']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['typical_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Common Procedures Same
    ws[f'A{current_row}'] = "Top 10 Common Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['common_same']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['common_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common procedures found"
        current_row += 1
    current_row += 2
    
    # Diagnoses - Outlier
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Diagnoses - Typical Same
    ws[f'A{current_row}'] = f"{selected_row['typical_same_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['typical_same']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['typical_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Common Diagnoses Same
    ws[f'A{current_row}'] = "Top 10 Common Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['common_same']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['common_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common diagnoses found"
        current_row += 1
    current_row += 2
    
    # Demographics Same
    ws[f'A{current_row}'] = "Demographic Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['demographics']['comparison_same']) > 0:
        for r in dataframe_to_rows(analysis_data['demographics']['comparison_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 2
    
    # Cost Same
    ws[f'A{current_row}'] = "Cost Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['costs']['comparison_same']) > 0:
        for r in dataframe_to_rows(analysis_data['costs']['comparison_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 3
    
    # CROSS-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"CROSS-GROUP ANALYSIS: {selected_row['outlier_label'].title()} vs {selected_row['overlap_label'].title()}"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Procedures Cross - Outlier
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Procedures Cross - Typical Other
    ws[f'A{current_row}'] = f"{selected_row['typical_other_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['typical_other']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['typical_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Common Procedures Other
    ws[f'A{current_row}'] = "Top 10 Common Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['procedures']['common_other']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['common_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common procedures found"
        current_row += 1
    current_row += 2
    
    # Diagnoses Cross - Outlier
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Diagnoses Cross - Typical Other
    ws[f'A{current_row}'] = f"{selected_row['typical_other_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['typical_other']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['typical_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 1
    
    # Common Diagnoses Other
    ws[f'A{current_row}'] = "Top 10 Common Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['diagnoses']['common_other']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['common_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common diagnoses found"
        current_row += 1
    current_row += 2
    
    # Demographics Other
    ws[f'A{current_row}'] = "Demographic Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    if len(analysis_data['demographics']['comparison_other']) > 0:
        for r in dataframe_to_rows(analysis_data['demographics']['comparison_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 2
    
    # Cost Distribution Analysis Cross-Group
    ws[f'A{current_row}'] = "Cost Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['costs']['comparison_other']) > 0:
        for r in dataframe_to_rows(analysis_data['costs']['comparison_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
            current_row += 1
    current_row += 2
    
    # === ADD CHARTS ===
    emb_path = create_embedding_comparison_chart(
        selected_row['outlier_pin'],
        selected_row['typical_same_pin'],
        selected_row['typical_other_pin'],
        selected_row['outlier_name'],
        selected_row['typical_same_name'],
        selected_row['typical_other_name']
    )
    try:
        ws.add_image(Image(emb_path), f"A{current_row}")
        os.remove(emb_path)  # Clean up temp file
    except Exception as e:
        ws[f'A{current_row}'] = f"Embedding chart error: {str(e)}"
    current_row += 25
    
    pca_path = create_pca_chart(
        selected_row['outlier_pin'],
        selected_row['typical_same_pin'],
        selected_row['typical_other_pin'],
        selected_row['outlier_label'],
        selected_row['overlap_label'],
        selected_row['outlier_name'],
        selected_row['typical_same_name'],
        selected_row['typical_other_name']
    )
    try:
        ws.add_image(Image(pca_path), f"A{current_row}")
        os.remove(pca_path)  # Clean up temp file
    except Exception as e:
        ws[f'A{current_row}'] = f"PCA chart error: {str(e)}"
    current_row += 25
    # === END ADD ===
    
    # Memory cleanup
    del analysis_data
    gc.collect()

def create_excel_export_optimized(limit=EXPORT_LIMIT):
    """Optimized Excel export with pre-computation and incremental saves."""
    
    print("Starting optimized Excel export...")
    
    # PRE-COMPUTE ALL DATA (Major performance boost)
    export_cases, precomputed_data = precompute_all_analysis_data()
    summary_stats, navigation_df = prepare_export_data(export_cases)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Tab 1: Summary Statistics
    print("Creating summary tab...")
    ws_summary = wb.create_sheet("Summary_Statistics")
    ws_summary['A1'] = "OVERLAP ANALYSIS SUMMARY"
    for r in dataframe_to_rows(summary_stats, index=False, header=True):
        ws_summary.append(r)
    style_worksheet(ws_summary, "OVERLAP ANALYSIS SUMMARY")
    
    # Tab 2: Navigation Index
    print("Creating navigation tab...")
    ws_nav = wb.create_sheet("Navigation_Index")
    ws_nav['A1'] = "OVERLAPPING PROVIDER ANALYSIS INDEX"
    ws_nav['A1'].font = Font(size=16, bold=True)
    
    # Add navigation headers
    headers = ['Case #', 'Provider Name', 'Label A', 'Label B', 'Overlap Score', 'Analysis Link']
    for col, header in enumerate(headers, 1):
        ws_nav.cell(row=3, column=col, value=header).font = Font(bold=True)
    
    # Add navigation data with hyperlinks
    for idx, row in navigation_df.iterrows():
        ws_nav[f'A{idx+4}'] = row['Case_Number']
        ws_nav[f'B{idx+4}'] = row['Provider_Name']
        ws_nav[f'C{idx+4}'] = row['Label_A']
        ws_nav[f