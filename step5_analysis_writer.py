from openpyxl.drawing.image import Image

def add_back_link(ws, tab_name):
    """Add back to index link."""
    ws['A1'] = '← BACK TO INDEX'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, italic=True, color='0000FF', underline='single')
    ws['A1'].hyperlink = "#Navigation_Index!A1"

def add_data_table(ws, data, start_row, title):
    """Add a data table to worksheet."""
    if len(data) > 0:
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(bold=True)
        start_row += 1
        
        for r in dataframe_to_rows(data, index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=start_row, column=c, value=value)
            start_row += 1
        start_row += 1
    else:
        ws[f'A{start_row}'] = f"{title}: No data available"
        start_row += 2
    
    return start_row

def write_single_analysis(ws, selected_row, analysis_data, tab_name):
    """Write complete analysis for one provider case."""
    
    # Add back link
    add_back_link(ws, tab_name)
    current_row = 3
    
    # Header information
    ws[f'A{current_row}'] = f"OVERLAPPING PROVIDER ANALYSIS: {selected_row['outlier_name']}"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    current_row += 2
    
    # Provider descriptions
    ws[f'A{current_row}'] = f"Selected Provider for Analysis: {selected_row['outlier_name']} ({selected_row['outlier_label'].title()})"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider embeddings indicate the provider is close to {selected_row['overlap_label'].title()}."
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['overlap_label'].title()} is {selected_row['typical_other_name']}"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['outlier_label'].title()} is {selected_row['typical_same_name']}"
    current_row += 2
    
    # Distance Metrics
    ws[f'A{current_row}'] = "Distance Metrics:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to own centroid ({selected_row['outlier_label'].title()}): {selected_row['own_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to overlapping centroid ({selected_row['overlap_label'].title()}): {selected_row['other_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Overlap coefficient: {selected_row['overlap_score']:.4f}"
    current_row += 3
    
    # WITHIN-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"WITHIN-GROUP ANALYSIS: {selected_row['outlier_label'].title()} Providers"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Procedures
    current_row = add_data_table(ws, analysis_data['procedures']['outlier'], current_row, 
                                f"{selected_row['outlier_name']} - Top 10 Procedures")
    current_row = add_data_table(ws, analysis_data['procedures']['typical_same'], current_row,
                                f"{selected_row['typical_same_name']} - Top 10 Procedures")
    current_row = add_data_table(ws, analysis_data['procedures']['common_same'], current_row,
                                "Top 10 Common Procedures")
    
    # Diagnoses
    current_row = add_data_table(ws, analysis_data['diagnoses']['outlier'], current_row,
                                f"{selected_row['outlier_name']} - Top 10 Diagnoses")
    current_row = add_data_table(ws, analysis_data['diagnoses']['typical_same'], current_row,
                                f"{selected_row['typical_same_name']} - Top 10 Diagnoses")
    current_row = add_data_table(ws, analysis_data['diagnoses']['common_same'], current_row,
                                "Top 10 Common Diagnoses")
    
    # Demographics and Costs
    current_row = add_data_table(ws, analysis_data['demographics']['comparison_same'], current_row,
                                "Demographic Distribution Analysis")
    current_row = add_data_table(ws, analysis_data['costs']['comparison_same'], current_row,
                                "Cost Distribution Analysis")
    
    # CROSS-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"CROSS-GROUP ANALYSIS: {selected_row['outlier_label'].title()} vs {selected_row['overlap_label'].title()}"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Cross-group procedures
    current_row = add_data_table(ws, analysis_data['procedures']['outlier'], current_row,
                                f"{selected_row['outlier_name']} - Top 10 Procedures")
    current_row = add_data_table(ws, analysis_data['procedures']['typical_other'], current_row,
                                f"{selected_row['typical_other_name']} - Top 10 Procedures")
    current_row = add_data_table(ws, analysis_data['procedures']['common_other'], current_row,
                                "Top 10 Common Procedures")
    
    # Cross-group diagnoses
    current_row = add_data_table(ws, analysis_data['diagnoses']['outlier'], current_row,
                                f"{selected_row['outlier_name']} - Top 10 Diagnoses")
    current_row = add_data_table(ws, analysis_data['diagnoses']['typical_other'], current_row,
                                f"{selected_row['typical_other_name']} - Top 10 Diagnoses")
    current_row = add_data_table(ws, analysis_data['diagnoses']['common_other'], current_row,
                                "Top 10 Common Diagnoses")
    
    # Cross-group demographics and costs
    current_row = add_data_table(ws, analysis_data['demographics']['comparison_other'], current_row,
                                "Demographic Distribution Analysis")
    current_row = add_data_table(ws, analysis_data['costs']['comparison_other'], current_row,
                                "Cost Distribution Analysis")
    
    # ADD CHARTS
    try:
        # Embedding chart
        emb_path = create_embedding_comparison_chart(
            selected_row['outlier_pin'], selected_row['typical_same_pin'], selected_row['typical_other_pin'],
            selected_row['outlier_name'], selected_row['typical_same_name'], selected_row['typical_other_name']
        )
        ws.add_image(Image(emb_path), f"A{current_row}")
        os.remove(emb_path)
        current_row += 25
        
        # PCA chart
        pca_path = create_pca_chart(
            selected_row['outlier_pin'], selected_row['typical_same_pin'], selected_row['typical_other_pin'],
            selected_row['outlier_label'], selected_row['overlap_label'],
            selected_row['outlier_name'], selected_row['typical_same_name'], selected_row['typical_other_name']
        )
        ws.add_image(Image(pca_path), f"A{current_row}")
        os.remove(pca_path)
        current_row += 25
        
    except Exception as e:
        ws[f'A{current_row}'] = f"Chart generation error: {str(e)}"
        current_row += 2
    
    # Memory cleanup
    del analysis_data
    gc.collect()

print("✅ Step 5 Complete: Analysis writer functions ready")