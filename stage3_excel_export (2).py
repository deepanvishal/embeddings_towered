import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime
import gc

# Configuration - Easy to change for testing vs full run
EXPORT_LIMIT = 5  # Change to None for all providers, or any number for testing
EXPORT_FILENAME = f"Overlapping_Provider_Analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

print(f"Excel Export Configuration:")
print(f"Export Limit: {EXPORT_LIMIT if EXPORT_LIMIT else 'All providers'}")
print(f"Output File: {EXPORT_FILENAME}")

def prepare_export_data(limit=None):
    """Prepare analysis data for export with optional limit for testing."""
    
    # Select top overlap cases by score
    export_cases = analysis_df.nlargest(limit if limit else len(analysis_df), 'overlap_score').copy()
    
    print(f"Preparing export data for {len(export_cases)} cases...")
    
    # Create summary statistics
    summary_stats = overlap_summary.copy()
    
    # Create navigation index
    navigation_index = []
    for idx, (_, row) in enumerate(export_cases.iterrows(), 1):
        navigation_index.append({
            'Case_Number': idx,
            'Provider_Name': row['outlier_name'],
            'Label_A': row['outlier_label'].title(),
            'Label_B': row['overlap_label'].title(), 
            'Overlap_Score': round(row['overlap_score'], 3),
            'Tab_Name': f"Case_{idx}_{row['outlier_name'][:20].replace(' ', '_')}"
        })
    
    navigation_df = pd.DataFrame(navigation_index)
    
    return summary_stats, navigation_df, export_cases

def get_analysis_data(selected_row):
    """Generate complete analysis data for one provider case."""
    
    # Basic info
    outlier_pin = selected_row['outlier_pin']
    typical_same_pin = selected_row['typical_same_pin'] 
    typical_other_pin = selected_row['typical_other_pin']
    
    # Get all comparison data
    proc_outlier = get_provider_codes(outlier_pin, 'Procedure', 10)
    proc_typical_same = get_provider_codes(typical_same_pin, 'Procedure', 10)
    proc_typical_other = get_provider_codes(typical_other_pin, 'Procedure', 10)
    
    diag_outlier = get_provider_codes(outlier_pin, 'Diagnosis', 10)
    diag_typical_same = get_provider_codes(typical_same_pin, 'Diagnosis', 10)
    diag_typical_other = get_provider_codes(typical_other_pin, 'Diagnosis', 10)
    
    common_proc_same = find_common_codes(outlier_pin, typical_same_pin, 'Procedure', 10)
    common_proc_other = find_common_codes(outlier_pin, typical_other_pin, 'Procedure', 10)
    
    common_diag_same = find_common_codes(outlier_pin, typical_same_pin, 'Diagnosis', 10)
    common_diag_other = find_common_codes(outlier_pin, typical_other_pin, 'Diagnosis', 10)
    
    demo_outlier = get_provider_demographics(outlier_pin)
    demo_typical_same = get_provider_demographics(typical_same_pin)
    demo_typical_other = get_provider_demographics(typical_other_pin)
    
    cost_outlier = get_provider_costs(outlier_pin)
    cost_typical_same = get_provider_costs(typical_same_pin)
    cost_typical_other = get_provider_costs(typical_other_pin)
    
    demo_comp_same = create_comparison_table(demo_outlier, demo_typical_same, 
                                           selected_row['outlier_name'], selected_row['typical_same_name'], 'demo')
    demo_comp_other = create_comparison_table(demo_outlier, demo_typical_other,
                                            selected_row['outlier_name'], selected_row['typical_other_name'], 'demo')
    
    cost_comp_same = create_comparison_table(cost_outlier, cost_typical_same,
                                           selected_row['outlier_name'], selected_row['typical_same_name'], 'cost') 
    cost_comp_other = create_comparison_table(cost_outlier, cost_typical_other,
                                            selected_row['outlier_name'], selected_row['typical_other_name'], 'cost')
    
    return {
        'procedures': {
            'outlier': proc_outlier,
            'typical_same': proc_typical_same,
            'typical_other': proc_typical_other,
            'common_same': common_proc_same,
            'common_other': common_proc_other
        },
        'diagnoses': {
            'outlier': diag_outlier,
            'typical_same': diag_typical_same, 
            'typical_other': diag_typical_other,
            'common_same': common_diag_same,
            'common_other': common_diag_other
        },
        'demographics': {
            'comparison_same': demo_comp_same,
            'comparison_other': demo_comp_other
        },
        'costs': {
            'comparison_same': cost_comp_same,
            'comparison_other': cost_comp_other
        }
    }

def style_worksheet(ws, title):
    """Apply professional styling to worksheet."""
    
    # Header styling
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Title styling  
    title_font = Font(name='Arial', size=16, bold=True, color='1F497D')
    
    # Set title
    ws['A1'] = title
    ws['A1'].font = title_font
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_hyperlinks_to_navigation(ws, navigation_df):
    """Add hyperlinks from navigation table to individual analysis tabs."""
    
    # Add hyperlinks in Tab_Link column
    for idx, row in navigation_df.iterrows():
        cell_ref = f"F{idx + 4}"  # Assuming hyperlinks go in column F, starting row 4
        tab_name = row['Tab_Name']
        ws[cell_ref] = f"→ {tab_name}"
        ws[cell_ref].hyperlink = f"#{tab_name}!A1"
        ws[cell_ref].font = Font(color='0000FF', underline='single')

def add_back_link(ws, tab_name):
    """Add back to index link at top of analysis tabs."""
    
    ws['A1'] = '← BACK TO INDEX'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, italic=True, color='0000FF', underline='single')
    ws['A1'].hyperlink = "#Navigation_Index!A1"

def write_analysis_to_worksheet(ws, selected_row, analysis_data, tab_name):
    """Write complete analysis data to worksheet with all sections."""
    
    # Add back link
    add_back_link(ws, tab_name)
    
    current_row = 3
    
    # Header information
    ws[f'A{current_row}'] = f"OVERLAPPING PROVIDER ANALYSIS: {selected_row['outlier_name']}"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    current_row += 2
    
    ws[f'A{current_row}'] = f"Selected Provider for Analysis: {selected_row['outlier_name']} ({selected_row['outlier_label'].title()})"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider embeddings indicate the provider is close to {selected_row['overlap_label'].title()}."
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['overlap_label'].title()} is {selected_row['typical_other_name']}"
    current_row += 1
    ws[f'A{current_row}'] = f"The provider closest to centroid of {selected_row['outlier_label'].title()} is {selected_row['typical_same_name']}"
    current_row += 2
    
    # Distance Metrics
    ws[f'A{current_row}'] = "Distance Metrics:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to own centroid ({selected_row['outlier_label'].title()}): {selected_row['own_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Distance to overlapping centroid ({selected_row['overlap_label'].title()}): {selected_row['other_distance']:.4f}"
    current_row += 1
    ws[f'A{current_row}'] = f"  Overlap coefficient: {selected_row['overlap_score']:.4f}"
    current_row += 3
    
    # WITHIN-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"WITHIN-GROUP ANALYSIS: {selected_row['outlier_label'].title()} Providers"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Procedure Volume Analysis
    ws[f'A{current_row}'] = "Procedure Volume Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2
    
    # Outlier Procedures
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['outlier']):  # Header row
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No procedure data available"
        current_row += 1
    current_row += 1
    
    # Typical Same Procedures
    ws[f'A{current_row}'] = f"{selected_row['typical_same_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['typical_same']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['typical_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['typical_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No procedure data available"
        current_row += 1
    current_row += 1
    
    # Common Procedures Within Group
    ws[f'A{current_row}'] = "Top 10 Common Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['common_same']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['common_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['common_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common procedures found in top sets"
        current_row += 1
    current_row += 2
    
    # Diagnosis Volume Analysis
    ws[f'A{current_row}'] = "Diagnosis Volume Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2
    
    # Outlier Diagnoses
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['outlier']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No diagnosis data available"
        current_row += 1
    current_row += 1
    
    # Typical Same Diagnoses
    ws[f'A{current_row}'] = f"{selected_row['typical_same_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['typical_same']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['typical_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['typical_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No diagnosis data available"
        current_row += 1
    current_row += 1
    
    # Common Diagnoses Within Group
    ws[f'A{current_row}'] = "Top 10 Common Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['common_same']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['common_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['common_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common diagnoses found in top sets"
        current_row += 1
    current_row += 2
    
    # Demographics Comparison Within Group
    ws[f'A{current_row}'] = "Demographic Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['demographics']['comparison_same']) > 0:
        for r in dataframe_to_rows(analysis_data['demographics']['comparison_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['demographics']['comparison_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    current_row += 2
    
    # Cost Distribution Analysis Within Group
    ws[f'A{current_row}'] = "Cost Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['costs']['comparison_same']) > 0:
        for r in dataframe_to_rows(analysis_data['costs']['comparison_same'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['costs']['comparison_same']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    current_row += 3
    
    # CROSS-GROUP ANALYSIS
    ws[f'A{current_row}'] = f"CROSS-GROUP ANALYSIS: {selected_row['outlier_label'].title()} vs {selected_row['overlap_label'].title()}"
    ws[f'A{current_row}'].font = Font(size=12, bold=True, color='1F497D')
    current_row += 2
    
    # Procedure Volume Analysis Cross-Group
    ws[f'A{current_row}'] = "Procedure Volume Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2
    
    # Outlier Procedures (repeat for cross comparison)
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['outlier']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No procedure data available"
        current_row += 1
    current_row += 1
    
    # Typical Other Procedures
    ws[f'A{current_row}'] = f"{selected_row['typical_other_name']} - Top 10 Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['typical_other']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['typical_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['typical_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No procedure data available"
        current_row += 1
    current_row += 1
    
    # Common Procedures Cross-Group
    ws[f'A{current_row}'] = "Top 10 Common Procedures:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['procedures']['common_other']) > 0:
        for r in dataframe_to_rows(analysis_data['procedures']['common_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['procedures']['common_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common procedures found in top sets"
        current_row += 1
    current_row += 2
    
    # Diagnosis Volume Analysis Cross-Group
    ws[f'A{current_row}'] = "Diagnosis Volume Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2
    
    # Outlier Diagnoses (repeat for cross comparison)
    ws[f'A{current_row}'] = f"{selected_row['outlier_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['outlier']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['outlier'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['outlier']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No diagnosis data available"
        current_row += 1
    current_row += 1
    
    # Typical Other Diagnoses
    ws[f'A{current_row}'] = f"{selected_row['typical_other_name']} - Top 10 Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['typical_other']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['typical_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['typical_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No diagnosis data available"
        current_row += 1
    current_row += 1
    
    # Common Diagnoses Cross-Group
    ws[f'A{current_row}'] = "Top 10 Common Diagnoses:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['diagnoses']['common_other']) > 0:
        for r in dataframe_to_rows(analysis_data['diagnoses']['common_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['diagnoses']['common_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No common diagnoses found in top sets"
        current_row += 1
    current_row += 2
    
    # Demographics Comparison Cross-Group
    ws[f'A{current_row}'] = "Demographic Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['demographics']['comparison_other']) > 0:
        for r in dataframe_to_rows(analysis_data['demographics']['comparison_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['demographics']['comparison_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    current_row += 2
    
    # Cost Distribution Analysis Cross-Group
    ws[f'A{current_row}'] = "Cost Distribution Analysis:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    if len(analysis_data['costs']['comparison_other']) > 0:
        for r in dataframe_to_rows(analysis_data['costs']['comparison_other'], index=False, header=True):
            for c, value in enumerate(r, 1):
                ws.cell(row=current_row, column=c, value=value)
                if current_row == ws.max_row - len(analysis_data['costs']['comparison_other']):
                    ws.cell(row=current_row, column=c).font = Font(bold=True)
            current_row += 1
    
    # Memory cleanup
    del analysis_data
    gc.collect()

def create_excel_export(limit=EXPORT_LIMIT):
    """Main function to create complete Excel export."""
    
    print("Starting Excel export creation...")
    
    # Prepare data
    summary_stats, navigation_df, export_cases = prepare_export_data(limit)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Tab 1: Summary Statistics
    ws_summary = wb.create_sheet("Summary_Statistics")
    ws_summary['A1'] = "OVERLAP ANALYSIS SUMMARY"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    # Add summary data
    for r in dataframe_to_rows(summary_stats, index=False, header=True):
        ws_summary.append(r)
    
    style_worksheet(ws_summary, "OVERLAP ANALYSIS SUMMARY")
    
    # Tab 2: Navigation Index
    ws_nav = wb.create_sheet("Navigation_Index") 
    ws_nav['A1'] = "OVERLAPPING PROVIDER ANALYSIS INDEX"
    ws_nav['A1'].font = Font(size=16, bold=True)
    
    # Add navigation data
    ws_nav['A3'] = "Case #"
    ws_nav['B3'] = "Provider Name"
    ws_nav['C3'] = "Label A"
    ws_nav['D3'] = "Label B" 
    ws_nav['E3'] = "Overlap Score"
    ws_nav['F3'] = "Analysis Link"
    
    for idx, row in navigation_df.iterrows():
        ws_nav[f'A{idx+4}'] = row['Case_Number']
        ws_nav[f'B{idx+4}'] = row['Provider_Name']
        ws_nav[f'C{idx+4}'] = row['Label_A']
        ws_nav[f'D{idx+4}'] = row['Label_B']
        ws_nav[f'E{idx+4}'] = row['Overlap_Score']
        ws_nav[f'F{idx+4}'] = f"→ {row['Tab_Name']}"
        ws_nav[f'F{idx+4}'].hyperlink = f"#{row['Tab_Name']}!A1"
        ws_nav[f'F{idx+4}'].font = Font(color='0000FF', underline='single')
    
    style_worksheet(ws_nav, "PROVIDER ANALYSIS INDEX")
    
    # Tabs 3+: Individual Analysis
    for idx, (_, selected_row) in enumerate(export_cases.iterrows(), 1):
        print(f"Processing case {idx}/{len(export_cases)}: {selected_row['outlier_name']}")
        
        tab_name = navigation_df.iloc[idx-1]['Tab_Name']
        ws_analysis = wb.create_sheet(tab_name)
        
        # Get analysis data
        analysis_data = get_analysis_data(selected_row)
        
        # Write to worksheet
        write_analysis_to_worksheet(ws_analysis, selected_row, analysis_data, tab_name)
        
        # Memory cleanup
        del analysis_data
        gc.collect()
    
    # Save workbook
    wb.save(EXPORT_FILENAME)
    print(f"Excel export saved: {EXPORT_FILENAME}")
    print(f"Created {len(export_cases)} analysis tabs + 2 summary tabs")
    
    return EXPORT_FILENAME

# Execute export
print("Ready to create Excel export...")
print(f"Configuration: {EXPORT_LIMIT if EXPORT_LIMIT else 'All'} cases")
print("Run: create_excel_export() to generate the file")